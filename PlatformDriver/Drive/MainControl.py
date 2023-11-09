import pandas as pd
import random
import sys
from PyQt5 import QtCore
from PyQt5.QtChart import *
from PyQt5.QtCore import Qt, QTimer, QDateTime
from PyQt5.QtGui import *
from PyQt5.QtSerialPort import QSerialPort
from PyQt5.QtWidgets import *
from io import StringIO

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from PlatformDriver.Drive import ActionControl
from PlatformDriver.Drive.GUI.UITest import Ui_MainWindow
from PlatformDriver.Drive.GUI.UI_Data import globals
from PlatformDriver.Drive.GUI.UI_Menu_Func import menufunc_menuCurrentGroupSettings
from PlatformDriver.Drive.GUI.UI_Menu_Func import menufunc_menuPortSettings
from PlatformDriver.Drive.GUI.UI_Window_Func import windowfunc_Formula_List
from PlatformDriver.Drive.GUI.UI_Window_Func import windowfunc_Port
from PlatformDriver.Drive.GUI.UI_Window_Func import windowfunc_Setting_List
from PlatformDriver.Drive.GUI.UI_Window_Func import windowfunc_Work_Control
from PlatformDriver.Drive.GUI.UI_Window_Func import windowfunc_Working_List

global group_all
group_all = []


class MyWindow(QMainWindow, Ui_MainWindow):

    def __init__(self):
        super(MyWindow, self).__init__()

        # adjust screen resolution
        self.desktop = QApplication.desktop()
        self.screenRect = self.desktop.screenGeometry()
        self.screenHeight = self.screenRect.height()
        self.screenWidth = self.screenRect.width()

        self.resize(96, 70)

        self.serial_test = QSerialPort()
        self.setupUi(self)
        self.initUI()
        # self.serialPortSetting()
        self.initChartMU()
        self.initChartMD()

        self.buffer = StringIO()

        self.pushButton_FL_add.clicked.connect(windowfunc_Formula_List.buttonAddFL)
        self.pushButton_FL_rm.clicked.connect(windowfunc_Formula_List.buttonRmFL)
        self.tableWidget_FL.itemClicked.connect(lambda item: windowfunc_Formula_List.tableFLClick(self, item))
        self.comboBox_ChemicalReagents.currentIndexChanged.connect(windowfunc_Formula_List.chemPositionChange)

        self.pushButton_SL_add.clicked.connect(windowfunc_Setting_List.buttonAddSL)
        self.pushButton_SL_rm.clicked.connect(windowfunc_Setting_List.buttonRmSL)
        self.tableWidget_SL.itemClicked.connect(lambda item: windowfunc_Setting_List.tableSLClick(self, item))

        self.pushButton_WL_add.clicked.connect(windowfunc_Working_List.buttonAddWL)
        self.pushButton_WL_rm.clicked.connect(windowfunc_Working_List.buttonRmWL)
        self.pushButton_WL_Pre.clicked.connect(windowfunc_Working_List.buttonPreWL)
        self.pushButton_WL_Next.clicked.connect(windowfunc_Working_List.buttonNextWL)
        self.pushButton_WL_Overall.clicked.connect(self.buttonOverall)
        self.pushButton_WL_Each.clicked.connect(self.buttonEach)
        self.actionNew.triggered.connect(self.getNew)
        self.pushButton_EXPORT.clicked.connect(self.buttonExportWL)
        self.pushButton_IMPORT.clicked.connect(self.buttonImportWL)
        self.pushButton_SAVE.clicked.connect(self.saveWL)
        self.comboBox_Group.currentIndexChanged.connect(self.groupSwitch)
        self.actionClear_current_group.triggered.connect(self.menuClearCurrentGroup)
        self.actionSet_current_group_to_default.triggered.connect(self.menuSetCurrentGroupToDefault)
        self.actionCurrent_group_settings.triggered.connect(menufunc_menuCurrentGroupSettings.menuCurrentGroupSettings)
        self.actionPort_Settings.triggered.connect(menufunc_menuPortSettings.menuPortSettings)

        self.pushButton_MU1.clicked.connect(self.startTestMU)
        self.pushButton_MU2.clicked.connect(self.stopTestMU)

        self.pushButton_MD1.clicked.connect(self.startTestMD)
        self.pushButton_MD2.clicked.connect(self.pauseTestMD)
        self.pushButton_MD3.clicked.connect(self.stopTestMD)

        self.pushButton_Serial_Open.clicked.connect(windowfunc_Port.serialOpen)
        self.pushButton_Serial_Close.clicked.connect(windowfunc_Port.serialClose)
        self.pushButton_Serial_Send.clicked.connect(self.serialSendData)

        self.serial_test.readyRead.connect(self.serialReadData)

        self.pushButton_UPLOAD.clicked.connect(windowfunc_Work_Control.buttonUpload)
        self.pushButton_START.clicked.connect(ActionControl.buttonStart)

    def initUI(self):
        global group_position
        self.pushButton_WL_Each.setEnabled(False)
        self.pushButton_Serial_Close.setEnabled(False)
        self.pushButton_Serial_Send.setEnabled(False)
        # self.pushButton_
        self.actionNew.setShortcut("Ctrl+N")
        self.comboBox_Group.addItem('Group 1')
        self.lineEdit_WL_Position.setText('1')
        self.label_Total_Positions.setText('/' + str(len(globals.group_current)))

        self.timer_MU = QTimer()
        self.timer_MU.timeout.connect(self.update_Chart_MU)
        self.updating_MU = False

        self.timer_MD = QTimer()
        self.timer_MD.setInterval(500)
        self.variable_MD = 0
        self.time_points_MD = []

    def serialPortSetting(self):
        global selected_port
        # set port name
        print(selected_port)
        self.serial_test.setPortName(selected_port)
        # set Baud rate
        self.serial_test.setBaudRate(QSerialPort.Baud9600)
        # set data bits
        # self.serial_test.setDataBits(QSerialPort.Data8)
        # set parity
        # self.serial_test.setParity(QSerialPort.NoParity)
        # set stop bits
        # self.serial_test.setStopBits(QSerialPort.OneStop)
        # set flow control
        # self.serial_test.setFlowControl(QSerialPort.NoFlowControl)

        # open serial port
        # self.serial_test.open(QSerialPort.ReadWrite)

    def serialOpen(self):
        if not self.serial_test.isOpen():
            self.serial_test.open(QSerialPort.ReadWrite)
            self.pushButton_Serial_Open.setEnabled(False)
            self.pushButton_Serial_Close.setEnabled(True)
            self.pushButton_Serial_Send.setEnabled(True)

    def serialClose(self):
        if self.serial_test.isOpen():
            self.serial_test.close()
            self.pushButton_Serial_Open.setEnabled(True)
            self.pushButton_Serial_Close.setEnabled(False)
            self.pushButton_Serial_Send.setEnabled(False)

    def serialSendData(self):
        data_send = self.textEdit_Serial.toPlainText()
        if data_send:
            self.serial_test.write(data_send.encode())

    def serialReadData(self):
        '''data_read = self.serial_test.readLine()
        if data_read:
            text_read = data_read.data().decode()
            self.label_ActionFeedback.setText(text_read)
            print(text_read)'''
        while self.serial_test.canReadLine():
            line_read = self.serial_test.readLine().data().decode()
            if line_read.strip():
                self.label_ActionFeedback.setText(line_read)
                print(line_read)

    def initChartMU(self):
        # Material Usage Bar Chart
        ## create bar set
        self.bar_set_MU = QBarSet("Variables")
        self.bar_set_MU.append(0)
        self.bar_set_MU.append(0)
        self.bar_set_MU.append(0)
        self.bar_set_MU.append(0)
        self.bar_set_MU.append(0)
        ## create bar series
        self.bar_series_MU = QBarSeries()
        self.bar_series_MU.append(self.bar_set_MU)
        self.bar_series_MU.setBarWidth(0.6)
        self.bar_series_MU.setLabelsVisible(True)
        self.bar_series_MU.setLabelsPrecision(3)
        ## creat X axis
        self.axis_x_MU = QBarCategoryAxis()
        self.axis_x_MU.append(["1", "2", "3", "4", "5"])
        ## creat Y axis
        self.axis_y_MU = QValueAxis()
        self.axis_y_MU.setRange(0, 10)
        self.axis_y_MU.setLabelFormat("%.1f")
        self.axis_y_MU.setTickCount(6)
        ## create the bar chart
        self.bar_chart_MU = QChart()
        self.bar_chart_MU.addSeries(self.bar_series_MU)
        self.bar_chart_MU.setAxisX(self.axis_x_MU)
        self.bar_chart_MU.setAxisY(self.axis_y_MU)
        self.bar_series_MU.attachAxis(self.axis_x_MU)
        self.bar_series_MU.attachAxis(self.axis_y_MU)
        self.bar_chart_MU.setTitle("Material Usage")
        self.bar_chart_MU.legend().setVisible(False)
        self.bar_chart_MU.setAnimationOptions(QChart.SeriesAnimations)
        self.bar_chart_MU.setAnimationDuration(500)
        ## add into UI window
        self.chart_view_MU = QChartView(self.bar_chart_MU)
        self.chart_view_MU.setRenderHint(QPainter.Antialiasing)
        group_layout_MU = QVBoxLayout()
        group_layout_MU.addWidget(self.chart_view_MU)
        self.groupBox_MaterialUsage.setLayout(group_layout_MU)

    def initChartMD(self):
        # Measurement Data Line Chart
        ## create line_series
        self.line_series_MD = QLineSeries()
        ## create axis
        self.axis_x_MD = QDateTimeAxis()
        self.axis_y_MD = QValueAxis()
        self.axis_y_MD.setRange(0, 10)
        self.axis_y_MD.setLabelFormat("%.1f")
        self.axis_y_MD.setTickCount(6)
        ## create the line chart
        self.line_chart_MD = QChart()
        self.line_chart_MD.addSeries(self.line_series_MD)
        self.line_chart_MD.setAxisX(self.axis_x_MD)
        self.line_chart_MD.setAxisY(self.axis_y_MD)
        self.line_series_MD.attachAxis(self.axis_x_MD)
        self.line_series_MD.attachAxis(self.axis_y_MD)
        self.line_chart_MD.setTitle("Measurement Data")
        self.line_chart_MD.legend().setVisible(False)
        self.line_chart_MD.setAnimationOptions(QChart.SeriesAnimations)
        self.line_chart_MD.setAnimationDuration(500)
        ## add into UI window
        self.chart_view_MD = QChartView(self.line_chart_MD)
        # self.chart_view_MD.setRenderHint(QPainter.Antialiasing)
        group_layout_MD = QVBoxLayout()
        group_layout_MD.addWidget(self.chart_view_MD)
        self.groupBox_MeasurementData.setLayout(group_layout_MD)

    def build_table_overall_WL(self, group):
        # establish the overall table
        position_count = len(group)
        self.tableWidget_WL.setColumnCount(position_count)
        ## get the row count!!!
        maxCount_FL = 0
        maxCount_SL = 0
        for p in range(0, position_count):
            if group[p] != 'empty':
                itemCount = len(group[p])
                count_FL = int(group[p][itemCount - 2])
                count_SL = int(group[p][itemCount - 1])
                if count_FL > maxCount_FL:
                    maxCount_FL = count_FL
                if count_SL > maxCount_SL:
                    maxCount_SL = count_SL
        self.tableWidget_WL.setRowCount(maxCount_FL + maxCount_SL)
        # set names of columns and rows
        rowName = []
        for i in range(0, maxCount_FL):
            rowName.append("Formula" + str(i + 1))
        for j in range(0, maxCount_SL):
            rowName.append("Setting" + str(j + 1))
        columnName = []
        for k in range(0, position_count):
            columnName.append("Position" + str(k + 1))
        self.tableWidget_WL.setVerticalHeaderLabels(rowName)
        self.tableWidget_WL.setHorizontalHeaderLabels(columnName)
        # insert every position's content into the table
        for pp in range(0, position_count):
            thisPosition = group[pp]
            if thisPosition != 'empty':
                rowCount_FL = int(thisPosition[len(thisPosition) - 2])
                rowCount_SL = int(thisPosition[len(thisPosition) - 1])
                ## insert this position's content
                for kf in range(0, rowCount_FL):
                    item_insert = QTableWidgetItem(thisPosition[2 * kf] + '\n' + thisPosition[2 * kf + 1])
                    item_insert.setTextAlignment(Qt.AlignCenter)
                    self.tableWidget_WL.setItem(kf, pp, item_insert)
                for ks in range(0, rowCount_SL):
                    item_insert = QTableWidgetItem(
                        thisPosition[2 * rowCount_FL + 2 * ks] + '\n' + thisPosition[2 * rowCount_FL + 2 * ks + 1])
                    item_insert.setTextAlignment(Qt.AlignCenter)
                    # print(maxCount_FL + ks)
                    self.tableWidget_WL.setItem(maxCount_FL + ks, pp, item_insert)

        self.tableWidget_WL.resizeColumnsToContents()
        self.tableWidget_WL.resizeRowsToContents()
        for column in range(self.tableWidget_WL.columnCount()):
            column_width = self.tableWidget_WL.columnWidth(column)
            self.tableWidget_WL.setColumnWidth(column, column_width + 30)
        for row in range(self.tableWidget_WL.rowCount()):
            column_width = self.tableWidget_WL.columnWidth(column)
            self.tableWidget_WL.setColumnWidth(column, column_width + 10)

    def save_current_group(self):
        group_name_current = self.comboBox_Group.currentText()
        if group_name_current not in group_all:
            group_all.append(group_name_current)
            group_all.append(globals.group_current)
            self.label_ActionFeedback.setText("Successfully saved new group: " + group_name_current)
        elif group_name_current in group_all:
            group_index_current = group_all.index(group_name_current)
            group_all[group_index_current + 1] = globals.group_current
            self.label_ActionFeedback.setText("Successfully saved group: " + group_name_current)

    def buttonOverall(self, text):
        # global globals.group_current
        print(globals.group_current)
        # Clear the table
        self.clear_table_WL()
        if all(item == 'empty' for item in globals.group_current) is False:
            # establish the overall table
            print(globals.group_current)
            self.build_table_overall_WL(globals.group_current)
        # set the accessibility of the buttons
        self.pushButton_WL_Each.setEnabled(True)
        self.pushButton_WL_Overall.setEnabled(False)
        self.pushButton_WL_Pre.setEnabled(False)
        self.pushButton_WL_Next.setEnabled(False)
        self.lineEdit_WL_Position.setEnabled(False)
        self.pushButton_WL_add.setEnabled(False)
        self.pushButton_WL_rm.setEnabled(False)

    def buttonEach(self):
        # global globals.group_current
        # Clear the table
        self.clear_table_WL()
        # get the saved position contents
        # position_number_saved = int(self.lineEdit_WL_Position.text())
        # position_data_saved = globals.group_current[position_number_saved - 1]
        position_number_saved = 1
        position_data_saved = globals.group_current[position_number_saved - 1]
        self.lineEdit_WL_Position.setText(str(position_number_saved))
        if position_data_saved != 'empty':
            itemNum = len(position_data_saved)
            rowCount_FL = int(position_data_saved[itemNum - 2])
            rowCount_SL = int(position_data_saved[itemNum - 1])
            # establish the saved table
            self.tableWidget_WL.setColumnCount(rowCount_FL)
            self.tableWidget_WL.setRowCount(4 + rowCount_SL)
            windowfunc_Working_List.build_table_each_WL(rowCount_FL, rowCount_SL, position_data_saved)

        self.pushButton_WL_Each.setEnabled(False)
        self.pushButton_WL_Overall.setEnabled(True)
        self.pushButton_WL_Pre.setEnabled(True)
        self.pushButton_WL_Next.setEnabled(True)
        self.lineEdit_WL_Position.setEnabled(True)
        self.pushButton_WL_add.setEnabled(True)
        self.pushButton_WL_rm.setEnabled(True)

    def getNew(self):
        QMessageBox.information(None, "Warning", "Pending developing")

    def buttonExportWL(self):  # export the overall view of WL
        # global globals.group_current
        # pop up a dialog
        dialogExportWL = QDialog()
        dialogExportWL.setWindowTitle("Export Working List")
        layout = QVBoxLayout()
        label = QLabel("Set sheet Name:")
        lineEdit = QLineEdit()
        group_name_current = self.comboBox_Group.currentText()
        lineEdit.setText(group_name_current)
        checkBox = QCheckBox()
        checkBox.setText("Also change the name in the control UI")
        checkBox.setChecked(True)
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialogExportWL.accept)
        button_box.rejected.connect(dialogExportWL.reject)
        layout.addWidget(label)
        layout.addWidget(lineEdit)
        layout.addWidget(checkBox)
        layout.addWidget(button_box)
        dialogExportWL.setLayout(layout)

        if dialogExportWL.exec() == QDialog.Accepted:
            if checkBox.isChecked():
                index = self.comboBox_Group.currentIndex()
                group_name_new = lineEdit.text()
                self.comboBox_Group.setItemText(index, group_name_new)
            sheet_name_WL = lineEdit.text()
            sheet_name_WL = sheet_name_WL if sheet_name_WL else 'Sheet1'

            # get the location of the xlsx file
            filename, _ = QFileDialog.getSaveFileName(self, "Choose the xlsx file", "", "Excel files (*.xlsx)")

            # 加载已有的工作簿，如果不存在则创建一个新的工作簿
            try:
                wb = load_workbook(filename)
            except FileNotFoundError:
                wb = load_workbook(template='template.xlsx')

            # export the tablewidget to the xlsx file
            rows = self.tableWidget_WL.rowCount()
            cols = self.tableWidget_WL.columnCount()
            ## create a DataFrame for pandas to save the data of table
            df = pd.DataFrame(columns=[self.tableWidget_WL.horizontalHeaderItem(col).text()
                                       for col in range(cols)])
            ## copy the data from table_WL to DataFrame
            for row in range(rows):
                data = [str(self.tableWidget_WL.item(row, col).text())
                        if self.tableWidget_WL.item(row, col) is not None else ''
                        for col in range(cols)]
                df = pd.concat([df, pd.DataFrame([data], columns=df.columns)], ignore_index=True)

            ## Create a writer for xlsx
            writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')
            # writer = pd.ExcelWriter(filename, engine='xlsxwriter', mode='w')
            ## write the DataFrame in xlsx file
            df.to_excel(writer, sheet_name=sheet_name_WL, index=False)
            ## get the sheet in the xlsx file
            worksheet = writer.sheets[sheet_name_WL]
            ## adjust the width of columns
            for i, col in enumerate(df.columns):
                column_width = max(df[col].astype(str).map(len).max(), len(col))
                worksheet.column_dimensions[chr(65 + i)].width = column_width + 4
            ## Set the alignment of the cell
            for row in worksheet.iter_rows(min_row=2, min_col=1,
                                           max_row=worksheet.max_row,
                                           max_col=worksheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(vertical='center', wrap_text=True)

            ## close the writer and save xlsx file
            writer.close()

            '''rows = len(globals.group_current)
            cols = self.tableWidget_WL.columnCount()
            data = []
            for row in range(rows):
                row_data = []
                for col in range(cols):
                    item = self.tableWidget_WL.item(row, col)
                    if item is None:
                        row_data.append('')
                    else:
                        row_data.append(item.text().replace('\n', ' | '))
                data.append(row_data)
            df = pd.DataFrame(data)

            with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
                df.to_excel(writer, sheet_name=sheet_name_WL, index=False, header=False)'''

    def buttonImportWL(self):
        global group_all
        # global globals.group_current
        # save the current group
        self.save_current_group()
        # select the xlsx file
        filename, _ = QFileDialog.getOpenFileName(None, "Open file", ".", "Excel files (*.xlsx)")
        if filename:
            try:
                sheets = pd.read_excel(filename, sheet_name=None).keys()
            except Exception as e:
                QMessageBox.warning(self, 'Error', 'Failed to read Excel file: {}'.format(str(e)))
            # pop up a choosing window
            dialogImportWL = QDialog()
            dialogImportWL.setWindowTitle("Select sheet")
            layout = QVBoxLayout()
            ## create label and combobox
            label = QLabel("Select a sheet:")
            combo = QComboBox()
            combo.addItems(sheets)
            ## create Ok and cancel buttons
            button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
            button_box.accepted.connect(dialogImportWL.accept)
            button_box.rejected.connect(dialogImportWL.reject)
            ## add widgets to layout
            layout.addWidget(label)
            layout.addWidget(combo)
            layout.addWidget(button_box)
            ## set layout
            dialogImportWL.setLayout(layout)

            if dialogImportWL.exec_() == QDialog.Accepted:
                chosen_sheet_name = combo.currentText()
                print(chosen_sheet_name)
                # get the chosen sheet
                df = pd.read_excel(filename, sheet_name=chosen_sheet_name)
                group_import = []
                for col in df.columns:
                    col_list = []
                    for cell in df[col]:
                        # skip if the cell is empty
                        if pd.isna(cell):
                            continue
                        # split if there is '|' and save separately
                        if isinstance(cell, str) and '|' in cell:
                            col_list.extend(cell.split('|'))
                        else:
                            col_list.append(cell)
                    numberFL = int(next((i for i, x in enumerate(col_list) if 'rpm' in x), -1) / 2)
                    numberSL = int((len(col_list) - numberFL * 2) / 2)
                    col_list.append(numberFL)
                    col_list.append(numberSL)
                    print(col_list)
                    group_import.append(col_list)
                # Clear the table
                self.clear_table_WL()
                # establish the overall table
                self.build_table_overall_WL(group_import)
                self.comboBox_Group.addItem(chosen_sheet_name)

                group_all.append(chosen_sheet_name)
                group_all.append(group_import)
                globals.group_current = group_import
                self.comboBox_Group.setCurrentText(chosen_sheet_name)
                self.label_ActionFeedback.setText("Successfully imported group: " + chosen_sheet_name)

    def saveWL(self):
        global group_all
        # global globals.group_current
        group_name_current = self.comboBox_Group.currentText()
        print(group_name_current)
        print(group_all)
        if group_name_current not in group_all:
            group_all.append(group_name_current)
            group_all.append(globals.group_current)
            self.label_ActionFeedback.setText("Successfully saved new group: " + group_name_current)
        elif group_name_current in group_all:
            dialogOverrideGroup = QDialog()
            dialogOverrideGroup.setWindowTitle("Confirmation")
            layout = QVBoxLayout()
            label = QLabel("Confirm overwrite the group info?")
            label2 = QLabel("Or save as a new group:")
            lineEdit = QLineEdit()
            group_name_current = self.comboBox_Group.currentText()
            lineEdit.setText(group_name_current + "_Edit")
            button_box = QDialogButtonBox()
            pushButton1 = QPushButton("Ovwewrite")
            pushButton2 = QPushButton("Save as new")
            pushButton3 = QPushButton("Cancel")
            button_box.addButton("Overwrite", QDialogButtonBox.AcceptRole)
            button_box.addButton("Cancel", QDialogButtonBox.RejectRole)
            button_box.addButton(pushButton2, QDialogButtonBox.ActionRole)
            button_box.accepted.connect(dialogOverrideGroup.accept)
            button_box.rejected.connect(dialogOverrideGroup.reject)

            def save_new():
                print("save as new button has been clicked")
                if lineEdit.text() in group_all:
                    QMessageBox.information(None, 'Failed', 'This name has existed.', QMessageBox.Ok)
                elif lineEdit.text() not in group_all:
                    group_all.append(lineEdit.text())
                    group_all.append(globals.group_current)
                    self.label_ActionFeedback.setText("Successfully saved as new group: " + lineEdit.text())
                    self.comboBox_Group.addItem(lineEdit.text())
                    dialogOverrideGroup.close()

            pushButton2.clicked.connect(save_new)
            layout.addWidget(label)
            layout.addWidget(label2)
            layout.addWidget(lineEdit)
            layout.addWidget(button_box)
            dialogOverrideGroup.setLayout(layout)

            if dialogOverrideGroup.exec() == QDialog.Accepted:
                print("overwrite button has been clicked")
                group_index_current = group_all.index(group_name_current)
                group_all[group_index_current + 1] = globals.group_current
                self.label_ActionFeedback.setText("Successfully saved group: " + group_name_current)

        print(group_all)

    def groupSwitch(self):
        global group_all
        # global globals.group_current
        '''# save the current group
        group_name_current = self.comboBox_Group.currentText()
        if group_name_current not in group_all:
            group_all.append(group_name_current)
            group_all.append(globals.group_current)
            self.label_ActionFeedback.setText("Successfully saved new group: " + group_name_current)
        elif group_name_current in group_all:
            group_index_current = group_all.index(group_name_current)
            group_all[group_index_current + 1] = globals.group_current
            self.label_ActionFeedback.setText("Successfully saved group: " + group_name_current)
        '''
        # switch the group
        group_name_switch = self.comboBox_Group.currentText()
        if group_name_switch in group_all:
            group_index_switch = group_all.index(group_name_switch)
            group_switch = group_all[group_index_switch + 1]
        self.label_ActionFeedback.setText("Switched to group: " + group_name_switch)
        # Clear the table
        self.clear_table_WL()
        if group_switch is not None:
            # establish the overall table
            self.build_table_overall_WL(group_switch)
        globals.group_current = group_switch
        self.label_Total_Positions.setText('/' + str(len(globals.group_current)))

    def menuClearCurrentGroup(self):
        # global globals.group_current
        globals.group_current = ['empty' for _ in globals.group_current]
        self.label_ActionFeedback.setText('Current group cleared!')

    def menuSetCurrentGroupToDefault(self):
        # global globals.group_current
        globals.group_current = ['empty'] * 12
        self.label_Total_Positions.setText('/' + str(len(globals.group_current)))
        self.label_ActionFeedback.setText('Current group set to default!')

    def startTestMU(self):
        if not self.updating_MU:
            self.updating_MU = True
            self.pushButton_MU1.setEnabled(False)
            self.pushButton_MU2.setEnabled(True)
            self.timer_MU.start(2000)
            self.label_ActionFeedback.setText("Start MU chart test")

    def stopTestMU(self):
        if self.updating_MU:
            self.updating_MU = False
            self.pushButton_MU1.setEnabled(True)
            self.pushButton_MU2.setEnabled(False)
            self.timer_MU.stop()
            self.label_ActionFeedback.setText("Stop MU chart test")

    def update_Chart_MU(self):
        # generate values of the 5 variables randomly
        values = [random.uniform(1, 10) for _ in range(5)]
        # update the set of the bar chart MU
        for i, value in enumerate(values):
            self.bar_set_MU.replace(i, value)
        # update X axis of the chart
        categories = [str(i + 1) for i in range(5)]
        self.axis_x_MU.setCategories(categories)

    def startTestMD(self):
        self.variable_MD = 0
        self.time_points_MD.clear()
        # Start the timer
        self.timer_MD.timeout.connect(self.update_Chart_MD)
        self.timer_MD.start()
        self.label_ActionFeedback.setText("Start MD chart test")

    def pauseTestMD(self):
        self.timer_MD.stop()
        self.label_ActionFeedback.setText("Pause MD chart test")

    def stopTestMD(self):
        self.timer_MD.stop()
        self.line_series_MD.clear()
        self.time_points_MD.clear()
        self.label_ActionFeedback.setText("Stop MD chart test")

    def update_Chart_MD(self):
        self.variable_MD = random.uniform(0, 10)
        current_time_MD = QDateTime.currentDateTime()

        self.time_points_MD.append(current_time_MD)
        self.line_series_MD.append(current_time_MD.toMSecsSinceEpoch(), self.variable_MD)

        min_time = current_time_MD.addSecs(-30)
        self.axis_x_MD.setRange(min_time, current_time_MD)


if __name__ == "__main__":
    if hasattr(QtCore.Qt, 'AA_EnableHighDpiScaling'):
        QApplication.setAttribute(QtCore.Qt.AA_DisableHighDpiScaling, True)
    # QtCore.QCoreApplication.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling, True)
    QtCore.QCoreApplication.setAttribute(QtCore.Qt.AA_UseHighDpiPixmaps, True)
    app = QApplication(sys.argv)
    win = MyWindow()
    win.show()
    sys.exit(app.exec_())
